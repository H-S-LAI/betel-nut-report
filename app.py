import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --- 1. 核心功能：修復亂碼並讀取資料 ---
def load_and_fix_csv(uploaded_file):
    """
    讀取上傳的 CSV 檔案，嘗試修復編碼問題 (Mojibake)。
    假設原始檔案是 Big5 編碼但被錯誤解讀。
    """
    try:
        # 讀取檔案內容為 bytes
        bytes_data = uploaded_file.getvalue()
        
        # 嘗試解碼策略：先用 Latin1 讀取 (保留原始字節)，再轉回 Big5
        # 這是處理 "雙重編碼" 常見的手法
        try:
            content = bytes_data.decode('utf-8') # 嘗試直接 utf-8
            # 如果內容看起來像亂碼 (例如包含 '©±'), 嘗試修復
            if '©±' in content:
                content = content.encode('latin1').decode('big5')
        except:
            # 如果 utf-8 失敗，嘗試直接 big5
             content = bytes_data.decode('big5')

        # 將字串轉為 DataFrame
        df = pd.read_csv(io.StringIO(content))
        
        # 欄位名稱對照表 (根據你的檔案分析結果)
        col_map = {
            '©±¦W': '店名',
            '«~¦W': '品名',
            '°â¶q': '售量'
        }
        
        # 嘗試重新命名欄位
        df = df.rename(columns=col_map)
        
        # 確保關鍵欄位存在
        if '店名' in df.columns and '售量' in df.columns:
            # 清理資料：去除空值，確保售量是數字
            df = df[['店名', '品名', '售量']].dropna()
            df['售量'] = pd.to_numeric(df['售量'], errors='coerce').fillna(0)
            return df
        else:
            return None
            
    except Exception as e:
        st.error(f"檔案 {uploaded_file.name} 讀取失敗: {e}")
        return None

# --- 2. 核心功能：處理 Excel 模板 ---
def fill_excel_template(template_path, combined_df, grains_per_pack_map):
    """
    讀取 Excel 模板，將數據填入對應的儲存格。
    不破壞原有格式。
    """
    # 載入 Workbook (使用 openpyxl)
    wb = load_workbook(template_path)
    ws = wb.active # 假設在第一個工作表

    # 1. 建立資料字典方便查詢: data_dict[店名][品名] = 售量
    data_dict = {}
    for index, row in combined_df.iterrows():
        store = str(row['店名']).strip()
        product = str(row['品名']).strip()
        sales = row['售量']
        
        if store not in data_dict:
            data_dict[store] = {}
        # 累加 (以防同一檔案有重複紀錄)
        data_dict[store][product] = data_dict[store].get(product, 0) + sales

    # 2. 解析 Excel 模板結構
    # 我們假設：
    # - 第 3 列 (Row 3) 是標題列，包含品名 (特幼, 幼大口...)
    # - 第 4 列開始 是店名 (A欄)
    
    # 掃描標題列 (找出每個品名在哪一欄)
    header_row = 3
    col_mapping = {} # {'特幼': 3, '幼大口': 6, ...} (Key是品名, Value是Column Index)
    
    # 注意：你的模板中有重複的 "品名" "售量" 標題
    # 我們需要看 "售量" 左邊那一格是什麼品名
    # 或者直接根據你的截圖，每個品名佔據兩欄 (品名, 售量) 或是特定的 merge cell
    # 為了準確，我們採取「掃描所有儲存格內容」的方式來定位品名
    
    # 更穩健的方法：直接看你的截圖結構
    # B欄: 特幼售量, F欄: 幼大口售量, H欄: 多粒售量... 
    # 讓我們用動態掃描第一列出現品名的地方，記錄它的 "售量" 應該填在哪 (通常是品名欄位 index + 1 或同一欄下)
    # 根據你的CSV輸出，模板結構如下：
    # Row 3: 店名 | 品名 | 售量 | 品名 | 售量 ...
    # 但 Row 4 (第一筆資料) 顯示： B欄是特幼售量, F欄是幼大口, H欄是多粒...
    # 這裡我們需要一個手動或半自動的 Mapping，或者依賴模板上的文字。
    
    # --- 掃描 Column 標題 ---
    # 為了避免讀到 "售量" 兩個字，我們紀錄具體的產品名稱位置
    product_col_index = {} # {'特幼': 3, '多粒': 8 ...} 這是指「售量」那一格的 Column Index
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        # 如果這一格有寫字，且不是 "店名" 或 "售量"，那它可能是品名標頭 (或者是合併儲存格的主標題)
        # 但根據你的 CSV 輸出，標題列很混亂。
        # 讓我們改用一個更聰明的方法：掃描第 4 列 (第一筆資料彰草店)，看哪裡有填數字，反推上面是什麼品名？
        # 不行，因為那是空的模板。
        
        # 替代方案：我們掃描 Row 3，只要找到品名文字，就假設它的 右邊一格 (或正下方) 是填寫數字的地方。
        # 根據你的檔案： B欄(index 2)下面填特幼數字? 不，根據 CSV preview:
        # Unnamed: 2 (C欄) 是特幼售量 (因為 B欄是 "特幼" 文字? 還是模板中 B 欄就是特幼售量?)
        
        # 讓我們直接用使用者上傳的那個 `檳榔銷售統計.xlsx` 裡面的 header 來對應
        # 假設 Row 3 的某些格子寫著 "特幼", "多粒"...
        if cell_value and isinstance(cell_value, str):
            clean_val = cell_value.strip()
            # 記錄品名所在的欄位。通常售量會填在同一欄(如果品名在header) 或是下一欄
            # 根據你的截圖，品名跟售量是分開的欄位。
            # 讓我們假設：如果讀到 "特幼"，那它的售量在 col + 1
            if clean_val in ['特幼', '多菁', '幼大口', '多粒', '多大口', '幼菁', '雙子星', '普通']: # 你的產品清單
               product_col_index[clean_val] = col + 1

    # 如果自動掃描失敗（例如模板標題是合併儲存格），我們提供一個 fallback 硬編碼 (根據你上傳的檔案內容)
    # 根據上傳檔案分析：
    # Col 3 (C): 特幼
    # Col 6 (F): 幼大口
    # Col 8 (H): 多粒
    # Col 10 (J): 多大口
    # Col 12 (L): 幼菁
    # Col 14 (N): 雙子星
    # 讓我們混合使用：先試著找，找不到就用預設位置
    if not product_col_index:
        # 注意：openpyxl 的 column index 從 1 開始
        # C=3, F=6, H=8, J=10, L=12, N=14
        # 但還缺 "多菁" 和 "普通"
        # 根據你的檔案 '北屯店' 那一行: 多菁在 Col 5 (E)
        product_col_index = {
            '特幼': 3,
            '多菁': 5,
            '幼大口': 6, # 修正
            '多粒': 8,
            '多大口': 10,
            '幼菁': 12,
            '雙子星': 14,
            '普通': 16 # 假設
        }

    # 3. 填寫數據
    # 記錄各產品的總和 (包數)
    total_sales_packs = {p: 0 for p in product_col_index}
    
    # 從第 4 列開始掃描店名
    for row in range(4, ws.max_row + 1):
        store_name_cell = ws.cell(row=row, column=1).value # A欄是店名
        if not store_name_cell:
            continue
            
        store_name = str(store_name_cell).strip()
        
        # 跳過 "銷售包數" 和 "銷售粒數" 這兩列，最後再算
        if "銷售" in store_name:
            continue
            
        if store_name in data_dict:
            for product, col_idx in product_col_index.items():
                if product in data_dict[store_name]:
                    val = data_dict[store_name][product]
                    ws.cell(row=row, column=col_idx).value = val
                    total_sales_packs[product] += val

    # 4. 填寫 "銷售包數" 和 "銷售粒數"
    # 我們需要找到這兩列在哪裡。通常在所有店名的下面。
    row_packs = None
    row_grains = None
    
    for row in range(4, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and "銷售包數" in str(val):
            row_packs = row
        if val and "銷售粒數" in str(val):
            row_grains = row

    if row_packs:
        for product, col_idx in product_col_index.items():
            # 填寫總包數
            ws.cell(row=row_packs, column=col_idx).value = total_sales_packs[product]
            
            # 填寫粒數設定 (綠色字) - 在包數的左邊一格 (col_idx - 1)
            # 只有當使用者有設定該產品的粒數時才填
            if product in grains_per_pack_map:
                grains_setting = grains_per_pack_map[product]
                # 寫入設定值
                cell_setting = ws.cell(row=row_packs, column=col_idx-1)
                cell_setting.value = grains_setting
                # 簡單設定一下樣式 (綠色字體需額外 import Font，這裡先略過以求簡潔)
                
    if row_grains:
        for product, col_idx in product_col_index.items():
            # 計算總粒數 = 總包數 * 設定的每包粒數
            if product in grains_per_pack_map:
                grains_setting = grains_per_pack_map[product]
                total_grains = total_sales_packs[product] * grains_setting
                ws.cell(row=row_grains, column=col_idx).value = total_grains

    # 存到記憶體中
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 3. Streamlit 介面 ---
st.set_page_config(page_title="檳榔報表生成器", layout="wide")
st.title("🏭 檳榔銷售報表自動生成")

st.markdown("### 1. 上傳檔案")
col1, col2 = st.columns(2)

with col1:
    template_file = st.file_uploader("上傳報表模板 (檳榔銷售統計.xlsx)", type=["xlsx"])
    
with col2:
    source_files = st.file_uploader("上傳原始數據 (特幼, 雙子星...)", type=["csv", "xls"], accept_multiple_files=True)

# 預設粒數設定
default_grains = {
    "特幼": 116,
    "幼大口": 50,
    "多粒": 146,
    "多大口": 137,
    "幼菁": 36,
    "雙子星": 0, # 假設
    "多菁": 0,   # 假設
    "普通": 0    # 假設
}

st.markdown("### 2. 設定參數 (每包粒數)")
st.info("請在此輸入每種產品的每包粒數，這會顯示在報表綠色字的位置，並用於計算銷售總粒數。")

# 動態產生輸入框
cols = st.columns(4)
user_grains_setting = {}

# 找出所有可能的產品名稱 (可以從預設列表，也可以從上傳的檔案中掃描，這裡先用預設清單)
products_list = list(default_grains.keys())

for i, product in enumerate(products_list):
    with cols[i % 4]:
        val = st.number_input(f"{product}", value=default_grains[product], step=1)
        user_grains_setting[product] = val

# --- 4. 執行按鈕 ---
if st.button("🚀 生成報表", type="primary"):
    if not template_file or not source_files:
        st.warning("請確保模板檔案與原始數據檔案都已上傳。")
    else:
        with st.spinner("正在處理亂碼與彙整數據..."):
            all_data = []
            
            # 讀取所有原始檔
            for f in source_files:
                df = load_and_fix_csv(f)
                if df is not None:
                    all_data.append(df)
            
            if all_data:
                combined_df = pd.concat(all_data, ignore_index=True)
                
                # 顯示預覽 (方便除錯)
                with st.expander("查看彙整後的數據 (預覽前 10 筆)"):
                    st.dataframe(combined_df.head(10))
                
                # 填入 Excel
                try:
                    result_excel = fill_excel_template(template_file, combined_df, user_grains_setting)
                    
                    st.success("報表生成成功！")
                    
                    # 下載按鈕
                    st.download_button(
                        label="📥 下載完成的報表 (Excel)",
                        data=result_excel,
                        file_name="已填寫_檳榔銷售統計.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"填寫 Excel 時發生錯誤: {e}")
            else:
                st.error("無法從上傳的檔案中讀取到有效數據，請檢查檔案格式。")